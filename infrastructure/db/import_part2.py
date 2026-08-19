"""
Import script for Part 2 Excel data into the AFACI database.
Adds 52 new products with their nutrients.

Run from afaci/ directory:
    python -m infrastructure.db.import_part2
"""

import asyncio
import uuid

import asyncpg
import openpyxl

DB_URL = "postgresql://erikomaraliev:d@localhost/afaci"
EXCEL_PATH = "../../data/Состав_part_2_25-05-2026.xlsx"

GREEN = "FF92D050"
BLUE = "FF00B0F0"

# Excel sheet name → DB region name
SHEET_TO_REGION = {
    "Иссык-кульская область": "Иссык-Кульский",
    "Ошская область": "Ошская область",
    "Таласская область": "Таласская область",
    "Нарынская область": "Нарынская область",
    "Чуйская область": "Чуйская область",
    "Джалал-Абадская область": "Джалал-Абадская область",
}


def get_db_category(excel_cat: str, product_name: str) -> str:
    p = product_name.lower()

    if excel_cat == "Мясной":
        if "рыба" in p:
            return "Рыба"
        if "курица" in p or "птиц" in p:
            return "Мясо птицы"
        return "Мясо"

    if excel_cat == "Молочный":
        return "Молоко и молочные продукты"

    if excel_cat == "Овощи, фрукты, ягоды и продукты их переработки (кроме соков)":
        fruits = [
            "яблок",
            "яблоки",
            "хурма",
            "гранат",
            "смородин",
            "клубник",
            "персик",
            "черешн",
            "курага",
            "кишмиш",
            "облепих",
            "чернослив",
            "сухофрукт",
        ]
        greens = ["шпинат", "листья"]
        if any(k in p for k in fruits):
            return "Фрукты, ягоды, сухофрукты"
        if any(k in p for k in greens):
            return "Зелень, травы, листья, салаты"
        return "Овощи и овощные продукты"

    if excel_cat == "Мука":
        return "Мука, продукты из муки"

    if excel_cat == "Сыпучий пищевой продукт":
        return "Сладости, кондитерские изделия"

    if excel_cat == "Масло":
        return "Жиры, масла"

    if excel_cat == "Зерновые":
        return "Крупы, злаки"

    if excel_cat == "Грибы":
        return "Грибы"

    if excel_cat == "Орехи":
        return "Орехи"

    return None


def parse_value(val) -> tuple[float | None, float | None]:
    """Parse '1,07±0,15' → (1.07, 0.15); '29.05' → (29.05, None); '-'/None → (None, None)."""
    if val is None:
        return None, None
    s = str(val).strip()
    if s in ("-", "", "н/д"):
        return None, None
    s = s.replace(",", ".")
    if "±" in s:
        parts = s.split("±")
        try:
            return float(parts[0].strip()), float(parts[1].strip())
        except ValueError:
            return None, None
    try:
        return float(s), None
    except ValueError:
        return None, None


# Column index → (nutrient_name_in_db, nutrient_type_name, unit_name)
# Columns 0=category, 1=product; skip section headers at 9, 28, 48+
COLUMN_MAP = {
    2: ("Массовая доля растворимых сухих веществ", "Химический состав", "%/100г"),
    # col 3: аскорбиновая кислота — not in DB, skip
    4: ("Зольность", "Химический состав", "%/100г"),
    5: ("Массовая доля влаги", "Химический состав", "%/100г"),
    6: ("Массовая доля белка", "Химический состав", "%/100г"),
    7: ("Массовая доля жира", "Химический состав", "%/100г"),
    8: ("Углеводы", "Химический состав", "%/100г"),
    # col 9: section header
    10: ("Валин", "Аминокислотный состав", "мг/100г"),
    11: ("Изолейцин", "Аминокислотный состав", "мг/100г"),
    12: ("Лейцин", "Аминокислотный состав", "мг/100г"),
    13: ("Лизин", "Аминокислотный состав", "мг/100г"),
    14: ("Метионин", "Аминокислотный состав", "мг/100г"),
    15: ("Треонин", "Аминокислотный состав", "мг/100г"),
    16: ("Триптофан", "Аминокислотный состав", "мг/100г"),
    17: ("Фенилаланин", "Аминокислотный состав", "мг/100г"),
    18: ("Аланин", "Аминокислотный состав", "мг/100г"),
    19: ("Аргинин", "Аминокислотный состав", "мг/100г"),
    20: ("Аспарагиновая кислота", "Аминокислотный состав", "мг/100г"),
    21: ("Гистидин", "Аминокислотный состав", "мг/100г"),
    22: ("Глицин", "Аминокислотный состав", "мг/100г"),
    23: ("Глутаминовая кислота", "Аминокислотный состав", "мг/100г"),
    24: ("Пролин", "Аминокислотный состав", "мг/100г"),
    25: ("Серин", "Аминокислотный состав", "мг/100г"),
    26: ("Тирозин", "Аминокислотный состав", "мг/100г"),
    27: ("Цистеин", "Аминокислотный состав", "мг/100г"),
    # col 28: section header
    29: ("Ca", "Минеральный состав", "мг/100г"),
    30: ("Na", "Минеральный состав", "мг/100г"),
    31: ("K", "Минеральный состав", "мг/100г"),
    32: ("P", "Минеральный состав", "мг/100г"),
    33: ("Mn", "Минеральный состав", "мг/100г"),
    34: ("Zn", "Минеральный состав", "мг/100г"),
    35: ("Se", "Минеральный состав", "мг/100г"),
    36: ("Cu", "Минеральный состав", "мг/100г"),
    37: ("Fe", "Минеральный состав", "мг/100г"),
    38: ("I", "Минеральный состав", "мг/100г"),
    39: ("B", "Минеральный состав", "мг/100г"),
    40: ("Li", "Минеральный состав", "мг/100г"),
    41: ("Al", "Минеральный состав", "мг/100г"),
    42: ("Mg", "Минеральный состав", "мг/100г"),
    43: ("V", "Минеральный состав", "мг/100г"),
    44: ("Ni", "Минеральный состав", "мг/100г"),
    45: ("Co", "Минеральный состав", "мг/100г"),
    46: ("Cr", "Минеральный состав", "мг/100г"),
    47: ("Sn", "Минеральный состав", "мг/100г"),
    # cols 48+: fatty acids — not in DB, skip
}


async def main():
    conn = await asyncpg.connect(DB_URL)

    # --- Load reference data from DB ---
    regions = {
        r["name"]: r["id"] for r in await conn.fetch("SELECT id, name FROM regions")
    }
    categories = {
        r["name"]: r["id"] for r in await conn.fetch("SELECT id, name FROM categories")
    }
    nutrient_names = {
        r["name"]: r["id"]
        for r in await conn.fetch("SELECT id, name FROM nutrients_names")
    }
    nutrient_types = {
        r["name"]: r["id"]
        for r in await conn.fetch("SELECT id, name FROM nutrients_types")
    }
    units = {r["name"]: r["id"] for r in await conn.fetch("SELECT id, name FROM units")}

    existing = {
        (r["region_id"], r["name"])
        for r in await conn.fetch("SELECT name, region_id FROM products")
    }

    wb = openpyxl.load_workbook(EXCEL_PATH)

    added_products = 0
    added_nutrients = 0
    skipped_exists = 0
    skipped_color = 0
    skipped_dup = 0
    errors = []

    seen_in_file = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        db_region_name = SHEET_TO_REGION[sheet_name]
        region_id = regions.get(db_region_name)
        if not region_id:
            errors.append(f"Region not found in DB: {db_region_name}")
            continue

        rows = list(ws.iter_rows(min_row=2))
        for row in rows:
            excel_cat = row[0].value
            product_name = row[1].value
            if not product_name or not excel_cat:
                continue

            product_name = product_name.strip()
            excel_cat = excel_cat.strip()

            fill = row[1].fill
            bg = fill.fgColor.rgb if fill.fgColor.type == "rgb" else "none"
            if bg in (GREEN, BLUE):
                skipped_color += 1
                continue

            file_key = (db_region_name, product_name)
            if file_key in seen_in_file:
                skipped_dup += 1
                continue
            seen_in_file.add(file_key)

            if (region_id, product_name) in existing:
                skipped_exists += 1
                continue

            db_cat_name = get_db_category(excel_cat, product_name)
            if not db_cat_name:
                errors.append(
                    f"Unknown category mapping: '{excel_cat}' / '{product_name}'"
                )
                continue
            category_id = categories.get(db_cat_name)
            if not category_id:
                errors.append(
                    f"Category not found in DB: '{db_cat_name}' (product: {product_name})"
                )
                continue

            product_id = uuid.uuid4()
            try:
                await conn.execute(
                    """INSERT INTO products (id, name, category_id, region_id, subcategory_id)
                       VALUES ($1, $2, $3, $4, NULL)""",
                    product_id,
                    product_name,
                    category_id,
                    region_id,
                )
                existing.add((region_id, product_name))
                added_products += 1
            except Exception as e:  # noqa: BLE001
                errors.append(
                    f"Failed to insert product '{product_name}' ({db_region_name}): {e}"
                )
                continue

            for col_idx, (nut_name, nut_type_name, unit_name) in COLUMN_MAP.items():
                cell_val = row[col_idx].value if col_idx < len(row) else None
                quantity, error_rate = parse_value(cell_val)
                if quantity is None:
                    continue

                name_id = nutrient_names.get(nut_name)
                type_id = nutrient_types.get(nut_type_name)
                unit_id = units.get(unit_name)

                if not name_id or not type_id or not unit_id:
                    errors.append(
                        f"Missing ref for nutrient '{nut_name}' (product: {product_name})"
                    )
                    continue

                try:
                    await conn.execute(
                        """INSERT INTO nutrients
                           (id, id_product, id_name_component, id_type_component, unit_id, quantity, error_rate)
                           VALUES ($1, $2, $3, $4, $5, $6, $7)
                           ON CONFLICT (id_product, id_name_component) DO UPDATE
                             SET quantity = EXCLUDED.quantity,
                                 error_rate = EXCLUDED.error_rate""",
                        uuid.uuid4(),
                        product_id,
                        name_id,
                        type_id,
                        unit_id,
                        quantity,
                        error_rate,
                    )
                    added_nutrients += 1
                except Exception as e:  # noqa: BLE001
                    errors.append(
                        f"Nutrient insert error '{nut_name}' / '{product_name}': {e}"
                    )

    await conn.close()

    print("=" * 60)
    print("ИМПОРТ ЗАВЕРШЁН")
    print("=" * 60)
    print(f"  Добавлено продуктов:  {added_products}")
    print(f"  Добавлено нутриентов: {added_nutrients}")
    print(f"  Пропущено (есть в БД):  {skipped_exists}")
    print(f"  Пропущено (цвет GREEN/BLUE): {skipped_color}")
    print(f"  Пропущено (дубликат в файле): {skipped_dup}")
    if errors:
        print(f"\n  Ошибки ({len(errors)}):")
        for e in errors:
            print(f"    - {e}")


if __name__ == "__main__":
    asyncio.run(main())
